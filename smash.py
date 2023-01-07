import openpyxl
import pandas as pd

def get_table_data(start_col, start_row, end_col, end_row, sheet): #~TEST
    data = []
    for row in sheet.iter_rows(min_col=start_col, min_row=start_row, max_col=end_col, max_row=end_row):
        print(row)
        if row[0].value is None: # if the first cell is empty, there is no more data
            break
        data.append([cell.value for cell in row])
    return data

# Give the location of the file 
path = r"C:\Users\Judah Wilson\OneDrive\Documents\smash.xlsx"
wb = openpyxl.load_workbook(path) 

squad_strike_df = pd.DataFrame(get_table_data(1, 1, 4, 100, wb['Squad Strike Table']))
new_header = squad_strike_df.iloc[0] #grab the first row for the header
squad_strike_df = squad_strike_df[1:] #take the data less the header row
squad_strike_df.columns = new_header #set the header row as the df header

squad_strike_sheet = wb['Squad Strike Table']

squad_strike_data = get_table_data(1, 2, 5, 100, squad_strike_sheet)

# Cell objects also have a row, column, 
# and coordinate attributes that provide 
# location information for the cell. 
  
# Note: The first row or 
# column integer is 1, not 0. 
  
# Cell object is created by using 
# sheet object's cell() method. 
cell = squad_strike_sheet.cell(row = 1, column = 1)
  
# Print value of cell object 
# using the value attribute 
print(cell.value) 